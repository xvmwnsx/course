from django.shortcuts import render, redirect,get_object_or_404
from .models import Schedule
from .forms import ScheduleForm
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote
from django.utils.translation import gettext as _
from datetime import timedelta
from django.utils.timezone import now

@login_required
def schedule_list(request): 
    user = request.user
    try:
        week_offset = int(request.GET.get("week_offset", 0))
    except ValueError:
        week_offset = 0

    today = now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=5) 

    if user.role == 'admin':
        schedules = Schedule.objects.filter(date__range=[start_of_week, end_of_week])
    elif user.role == 'teacher':
        schedules = Schedule.objects.filter(teacher=user, date__range=[start_of_week, end_of_week])
    elif user.role == 'student' and user.group:
        schedules = Schedule.objects.filter(subject__group=user.group, date__range=[start_of_week, end_of_week])
    else:
        schedules = Schedule.objects.none()
    
    week_days_translation = {
        "Monday": _("Понедельник"),
        "Tuesday": _("Вторник"),
        "Wednesday": _("Среда"),
        "Thursday": _("Четверг"),
        "Friday": _("Пятница"),
        "Saturday": _("Суббота"),
    }
    week_days = list(week_days_translation.values())


    grouped_schedules = {day: [] for day in week_days}
    for schedule in schedules:
        english_day = schedule.date.strftime("%A") 
        russian_day = week_days_translation.get(english_day, None)  
        if russian_day:
            grouped_schedules.setdefault(russian_day, []).append(schedule) 

    can_edit = user.role in ['teacher', 'admin']

    return render(request, 'schedule_list.html', {
        'schedules': grouped_schedules,
        'can_edit': can_edit,
        'week_days': week_days,
        'week_offset': week_offset,
        'prev_week_offset': week_offset - 1,
        'next_week_offset': week_offset + 1,
        'start_of_week': start_of_week,
        'end_of_week': end_of_week
    })

@login_required
def schedule_edit(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    if request.user.role not in ['teacher', 'admin']:
        return HttpResponseForbidden("У вас нет прав на редактирование расписания.")

    if request.method == "POST":
        form = ScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            return redirect('schedule_list')
    else:
        form = ScheduleForm(instance=schedule)

    return render(request, 'schedule_edit.html', {'form': form})

@login_required
def schedule_search(request):
    query = request.GET.get('q', '')
    results = None
    if query:
        results = Schedule.objects.filter(Q(subject__name__icontains=query) | Q(subject__group__name__icontains=query) ).select_related('subject')  

    return render(request, 'schedule_search.html', 
                  {'query': query, 'results': results})

def download_schedule(request):
    user = request.user

    try:
        week_offset = int(request.GET.get("week_offset", 0))
    except ValueError:
        week_offset = 0

    today = now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=5) 

    if user.role == 'admin':
        schedules = Schedule.objects.filter(date__range=[start_of_week, end_of_week])
    elif user.role == 'teacher':
        schedules = Schedule.objects.filter(teacher=user, date__range=[start_of_week, end_of_week])
    elif user.role == 'student' and user.group:
        schedules = Schedule.objects.filter(subject__group=user.group, date__range=[start_of_week, end_of_week])
    else:
        schedules = Schedule.objects.none()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Расписание"

    headers = ["ДЕНЬ НЕДЕЛИ", "ПРЕДМЕТ", 'НОМЕР ГРУППЫ', 'ГРУППА', "ДАТА", "ВРЕМЯ", "КАБИНЕТ", "ПРЕПОДАВАТЕЛЬ"]
    sheet.append(headers)

    header_font = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="5b74e3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Side(border_style="thin", color="000000")
    header_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    days_of_week = {
        "Monday": "Понедельник",
        "Tuesday": "Вторник",
        "Wednesday": "Среда",
        "Thursday": "Четверг",
        "Friday": "Пятница",
        "Saturday": "Суббота",
        "Sunday": "Воскресенье"
    }

    for schedule_item in schedules:
        teacher_name = f"{schedule_item.teacher.first_name} {schedule_item.teacher.last_name}"
        weekday = days_of_week.get(schedule_item.date.strftime('%A'), "Неизвестно")

        sheet.append([
            weekday,
            schedule_item.subject.name,
            schedule_item.subject.group.id,
            schedule_item.subject.group.name,
            schedule_item.date.strftime("%d-%m-%Y"),
            schedule_item.time.strftime("%H:%M"),
            schedule_item.cabinet or "—",
            teacher_name,
        ])

    data_alignment = Alignment(horizontal="left", vertical="center")
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = header_border

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "Расписание занятий.xlsx"
    response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{quote(filename)}'
    workbook.save(response)

    return response